#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
証憑画像認識システム v1.3統合版 - ポータブル版 v3 (複数商品対応版)
新しいExcel形式対応 + 画像整理機能 + 複数商品認識
"""

import os
import sys
import json
import hashlib
import base64
import shutil
import re
from pathlib import Path
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Union
import time
import logging

# サードパーティライブラリ
try:
    from openai import OpenAI
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from PIL import Image
    from dotenv import load_dotenv
except ImportError as e:
    print(f"必要なライブラリがインストールされていません: {e}")
    print("pip install openai openpyxl pillow python-dotenv を実行してください")
    input("Enterで終了...")
    sys.exit(1)

class PortableReceiptProcessorV3Multi:
    def __init__(self):
        # 実行環境の検出
        if getattr(sys, 'frozen', False):
            # PyInstaller実行ファイル
            executable_dir = Path(sys.executable).parent
            print(f"実行ファイルディレクトリ: {executable_dir}")
            self.base_dir = executable_dir
            
            if self.base_dir.name == "システムファイル":
                self.base_dir = self.base_dir.parent
        else:
            # Python スクリプト実行
            script_dir = Path(__file__).parent
            print(f"スクリプトディレクトリ: {script_dir}")
            
            if script_dir.name == "システムファイル":
                self.base_dir = script_dir.parent
            else:
                self.base_dir = script_dir
        
        print(f"ベースディレクトリ: {self.base_dir}")
        
        # フォルダパスの設定（仕様書に従い絵文字付きフォルダ名）
        self.image_folder = self.base_dir / "📁 画像フォルダ"
        self.output_folder = self.base_dir / "📊 出力ファイル"
        self.settings_folder = self.base_dir / "⚙️ 設定ファイル"
        
        # 処理済みファイル記録の設定
        self.processed_file = self.base_dir / "システムファイル" / "processed_images.json"
        if not self.processed_file.parent.exists():
            self.processed_file = self.base_dir / "processed_images.json"
        
        # 必要なフォルダを作成
        for folder in [self.image_folder, self.output_folder, self.settings_folder]:
            folder.mkdir(exist_ok=True)
        
        # 処理済み画像リストを読み込み
        self.processed_images = self.load_processed_images()
        
        # OpenAI クライアント
        self.openai_client = None

    def load_processed_images(self) -> set:
        """処理済み画像ハッシュを読み込み"""
        if self.processed_file.exists():
            try:
                with open(self.processed_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                return set(data.get('processed_hashes', []))
            except Exception as e:
                print(f"処理済みファイル読み込みエラー: {e}")
        return set()

    def save_processed_images(self):
        """処理済み画像ハッシュを保存"""
        try:
            data = {
                'processed_hashes': list(self.processed_images),
                'last_updated': datetime.now().isoformat()
            }
            with open(self.processed_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"処理済みファイル保存エラー: {e}")

    def archive_processed_images(self, processed_files: List[Path], results: List[Dict[str, Any]]):
        """処理済み画像をYYYYMM形式でアーカイブ"""
        try:
            print(f"\n📦 処理済み画像をアーカイブ中...")
            
            # ファイルと結果をマッピング
            file_result_map = {}
            for result in results:
                file_name = result.get('file_name')
                if file_name:
                    file_result_map[file_name] = result
            
            moved_count = 0
            current_year = datetime.now().year
            current_month = datetime.now().month
            
            for file_path in processed_files:
                if not file_path.exists():
                    print(f"⚠️ ファイルが見つかりません: {file_path.name}")
                    continue
                
                # 対応する結果を取得
                result = file_result_map.get(file_path.name)
                year_month = f"{current_year}{current_month:02d}"  # デフォルト: 現在年月
                
                if result and result.get('purchase_date'):
                    try:
                        date_parts = result['purchase_date'].split('/')
                        if len(date_parts) >= 2:
                            year = int(date_parts[0]) if len(date_parts) >= 3 else current_year
                            month = int(date_parts[1] if len(date_parts) >= 3 else date_parts[0])
                            if 1 <= month <= 12:
                                year_month = f"{year}{month:02d}"
                    except:
                        pass
                
                # アーカイブフォルダを作成
                archive_folder = self.output_folder / year_month
                archive_folder.mkdir(exist_ok=True)
                
                # ファイルを移動
                dest_path = archive_folder / file_path.name
                try:
                    shutil.move(str(file_path), str(dest_path))
                    print(f"📁 {file_path.name} → {year_month}/")
                    moved_count += 1
                except Exception as e:
                    print(f"⚠️ 移動エラー ({file_path.name}): {e}")
            
            print(f"✅ {moved_count}個の画像をアーカイブしました")
            
        except Exception as e:
            print(f"❌ アーカイブエラー: {e}")

    def validate_address_format(self, address_text: str) -> tuple:
        """住所の基本フォーマットを検証（郵便番号または都道府県開始）"""
        if not address_text or len(address_text.strip()) < 3:
            return False, "住所が短すぎます"
        
        address_text = address_text.strip()
        
        # 郵便番号チェック
        if re.match(r'^〒?\d{3}-?\d{4}', address_text):
            return True, "郵便番号形式で有効"
        
        # 都道府県チェック
        prefectures = [
            '北海道', '青森県', '岩手県', '宮城県', '秋田県', '山形県', '福島県',
            '茨城県', '栃木県', '群馬県', '埼玉県', '千葉県', '東京都', '神奈川県',
            '新潟県', '富山県', '石川県', '福井県', '山梨県', '長野県', '岐阜県',
            '静岡県', '愛知県', '三重県', '滋賀県', '京都府', '大阪府', '兵庫県',
            '奈良県', '和歌山県', '鳥取県', '島根県', '岡山県', '広島県', '山口県',
            '徳島県', '香川県', '愛媛県', '高知県', '福岡県', '佐賀県', '長崎県',
            '熊本県', '大分県', '宮崎県', '鹿児島県', '沖縄県'
        ]
        
        for prefecture in prefectures:
            if address_text.startswith(prefecture):
                return True, f"{prefecture}で開始する有効な住所"
        
        return False, "郵便番号または都道府県で始まらない無効な住所"

    def validate_invoice_number(self, invoice_number: str) -> str:
        """インボイス番号のフォーマット検証"""
        if not invoice_number:
            return ""
        
        # T + 13桁数字のパターンチェック
        pattern = r'^T\d{13}$'
        
        if re.match(pattern, invoice_number.strip()):
            return invoice_number.strip()
        else:
            print(f"⚠️ 無効なインボイス番号形式: {invoice_number} (T+13桁の形式ではありません)")
            return ""

    def search_shop_address_with_ai(self, shop_name: str) -> Dict[str, Any]:
        """OpenAI APIを使用した店舗住所検索（知識カットオフ明示方式）"""
        if not shop_name or not self.openai_client:
            return {"has_data": False, "address": "", "note": "検索条件不足"}
        
        try:
            print(f"🔍 AIで店舗住所を検索中: {shop_name}")
            
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[{
                    "role": "user",
                    "content": f"""あなたの学習データ内で「{shop_name}」の住所情報を以下のJSON形式で回答してください：

{{
  "shop_name": "{shop_name}",
  "address": "住所（学習データにある場合のみ）",
  "confidence": "確信度(1-10)",
  "has_data": true/false,
  "note": "補足説明"
}}

【重要ルール】:
- 学習データに含まれている場合のみ住所を記載
- 推測・想像での回答は禁止  
- 不明な場合は has_data: false を設定
- チェーン店の場合は「チェーン店のため特定不可」をnoteに記載
- 住所は都道府県から番地まで可能な限り詳細に"""
                }],
                max_tokens=500,
                temperature=0.1
            )
            
            content = response.choices[0].message.content
            print(f"🤖 AI応答: {content}")
            
            # JSON抽出
            json_match = re.search(r'\{.*\}', content, re.DOTALL)
            if json_match:
                result = json.loads(json_match.group())
                
                # 住所の基本検証
                if result.get('has_data') and result.get('address'):
                    is_valid, reason = self.validate_address_format(result['address'])
                    if not is_valid:
                        print(f"⚠️ AI検索住所が無効: {result['address']} → {reason}")
                        result['has_data'] = False
                        result['note'] = f"住所形式無効: {reason}"
                    else:
                        print(f"✅ AI検索成功: {result['address']} (確信度: {result.get('confidence', 'N/A')})")
                
                return result
            else:
                return {"has_data": False, "address": "", "note": "JSON解析エラー"}
                
        except Exception as e:
            print(f"❌ AI住所検索エラー: {e}")
            return {"has_data": False, "address": "", "note": f"検索エラー: {str(e)}"}

    def get_image_hash(self, image_path: Path) -> str:
        """画像ファイルのハッシュ値を計算"""
        try:
            with open(image_path, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        except Exception as e:
            print(f"ハッシュ計算エラー ({image_path.name}): {e}")
            return None

    def setup_api_key(self):
        """OpenAI APIキーの設定"""
        env_file = self.settings_folder / ".env"
        
        # .envファイルから読み込み
        if env_file.exists():
            load_dotenv(env_file)
            api_key = os.getenv("OPENAI_API_KEY")
            if api_key:
                try:
                    self.openai_client = OpenAI(api_key=api_key)
                    self.openai_client.models.list()
                    print("\n✅ OpenAI APIキーが正常に設定されました")
                    return True
                except Exception as e:
                    print(f"\n❌ APIキーテストに失敗: {e}")
        
        # 新しいAPIキーを入力
        print("\n🔑 OpenAI APIキーを設定してください")
        print("💡 APIキーの取得: https://platform.openai.com/api-keys")
        api_key = input("APIキーを入力 (sk-...): ").strip()
        
        if not api_key:
            print("❌ APIキーが入力されませんでした")
            return False
        
        # APIキーのテスト
        try:
            self.openai_client = OpenAI(api_key=api_key)
            self.openai_client.models.list()
            
            # 設定を保存
            with open(env_file, 'w', encoding='utf-8') as f:
                f.write(f"OPENAI_API_KEY={api_key}\n")
            
            print("✅ APIキーが正常に設定されました")
            return True
            
        except Exception as e:
            print(f"❌ APIキーが無効です: {e}")
            return False

    def get_new_images(self) -> List[Path]:
        """新しい画像ファイルを取得"""
        supported_extensions = ['.jpg', '.jpeg', '.png', '.heic', '.heif']
        new_images = []
        
        print(f"画像フォルダをスキャン中: {self.image_folder}")
        
        if not self.image_folder.exists():
            print(f"❌ 画像フォルダが見つかりません: {self.image_folder}")
            return []
        
        # 全ファイルをチェック
        all_files = list(self.image_folder.iterdir())
        print(f"📁 発見されたファイル数: {len(all_files)}")
        
        for file_path in all_files:
            print(f"🔍 チェック中: {file_path.name} (拡張子: {file_path.suffix.lower()})")
            
            if file_path.suffix.lower() in supported_extensions:
                image_hash = self.get_image_hash(file_path)
                if image_hash not in self.processed_images:
                    new_images.append(file_path)
                    print(f"📋 新しい画像として追加: {file_path.name}")
                else:
                    print(f"⏭ 処理済みをスキップ: {file_path.name}")
            else:
                print(f"🚫 対応外の形式をスキップ: {file_path.name}")
        
        return sorted(new_images)

    def optimize_image(self, image_path: Path) -> tuple:
        """画像の最適化とEXIF情報処理"""
        try:
            img = Image.open(image_path)
            
            # EXIF情報の処理
            try:
                exif = img._getexif()
                if exif is not None:
                    for orientation in exif:
                        if orientation == 274:  # Orientation tag
                            exif_orientation = exif[orientation]
                            if exif_orientation == 3:
                                img = img.rotate(180, expand=True)
                            elif exif_orientation == 6:
                                img = img.rotate(270, expand=True)
                            elif exif_orientation == 8:
                                img = img.rotate(90, expand=True)
            except:
                pass
            
            # サイズ調整
            max_dimension = 4096
            if max(img.size) > max_dimension:
                img.thumbnail((max_dimension, max_dimension), Image.Resampling.LANCZOS)
            
            return img, None
        except Exception as e:
            return None, str(e)

    def analyze_receipt_with_vision(self, image_path: Path, retry_count: int = 0) -> Dict[str, Any]:
        """GPT-4 Visionでレシートを解析"""
        try:
            # 画像を最適化
            img, error = self.optimize_image(image_path)
            if error:
                raise Exception(f"画像最適化エラー: {error}")
            
            # base64エンコード
            from io import BytesIO
            buffer = BytesIO()
            img.save(buffer, format='JPEG', quality=98)
            base64_image = base64.b64encode(buffer.getvalue()).decode('utf-8')
            
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": "あなたは日本語レシート・領収書専用の情報抽出AIです。複数商品がある場合は商品ごとに個別情報を抽出し、Excel出力で商品別の行を作成できるようにしてください。"
                    },
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": f"""この画像から情報を抽出し、以下のJSON形式で返してください。

重要: 送料が商品リスト内に記載されている場合の特別処理

抽出項目:
1. purchase_date: 購入日（YYYY/MM/DD形式）
2. products: 商品リスト（配列形式、各商品を個別管理）
   - name: 商品名（型番込みで完全な名称）
     ※送料項目の場合は元の送料名称を保持（「小計」「合計」等は使用禁止）
     ※例: 「チュウコカメラ ヨヤクソウリョウ」は「チュウコカメラ ヨヤクソウリョウ」のまま記録
   - price: 商品単価（その商品のみの価格）
   - quantity: 数量（デフォルト1）
   - is_shipping: 送料項目かどうか（true/false）
3. shipping_fee: 送料（別途記載の場合のみ）
4. total_amount: 合計金額（全商品+送料の総額）
5. shop_name: 店舗名・出品者名
6. shop_address: 店舗住所・出品者住所
   ■【重要】オークション・ECサイトの場合の住所抽出ルール:
   ■ 出品者情報の住所のみ抽出（落札者情報は無視）
   ■ 「お届け情報」「配送先」「送付先」「落札者情報」の住所は記録しない
   ■ 「出品者情報」「販売者情報」「店舗情報」の住所のみ記録
   ■ Yahoo!オークション等では出品者の住所を優先
   ■ お届け先住所しかない場合は空文字列を設定
7. payment_method: 支払方法
8. shop_type: 店舗タイプ（実店舗/ヤフオク/ECサイト/フリマアプリ/その他）
9. invoice_number: インボイス番号（適格請求書発行事業者登録番号）
   ■ 正確なフォーマット: T + 数字13桁（例: T1234567890123）
   ■ 「登録番号」「適格請求書発行事業者」「インボイス番号」等の表記を探す
   ■ T以外で始まる番号や桁数が異なる場合は記録しない
   ■ 見つからない場合は空文字列
10. file_name: ファイル名

送料特別処理ルール:
■ 送料が商品リスト内に記載されている場合:
  - 以下のキーワードを含む項目を送料として識別:
    ・日本語: 送料、配送料、配送費、運賃、運送料、予約送料
    ・カタカナ: ソウリョウ、ハイソウリョウ、ハイソウヒ、ウンチン、ウンソウリョウ、ヨヤクソウリョウ
    ・複合語: チュウコカメラ ヨヤクソウリョウ、中古カメラ予約送料
    ・英語: shipping、delivery、postage
    ・配送会社: 佐川急便送料、ヤマト運輸送料等
  - 送料項目には is_shipping: true を設定
  - 送料項目の price に送料金額を設定
  - 全体の shipping_fee は "0" に設定（商品リスト内処理のため）

■ 送料が別途記載されている場合:
  - products には含めず、shipping_fee に設定
  - 商品項目には is_shipping: false を設定

■ 特別な認識ルール:
  - 「小計」が送料文脈で使用されている場合は送料として判定
  - 例: 「チュウコカメラ ヨヤクソウリョウ」→「小計: 1650円」の場合
  - この「小計」は送料項目として is_shipping: true で処理
  - 重要: 商品名は「小計」ではなく、元の送料を表す名称を記録
    ・「チュウコカメラ ヨヤクソウリョウ」→ name: "チュウコカメラ ヨヤクソウリョウ"
    ・「ヨヤクソウリョウ」→ name: "ヨヤクソウリョウ" 
    ・「予約送料」→ name: "予約送料"
  - 「小計」「合計」等の計算用語は商品名として使用しない

商品価格の考え方:
- 通常商品: 各商品の個別価格を抽出、is_shipping: false
- 送料項目: 送料金額を price に設定、is_shipping: true
- 商品価格の合計 + 送料 = 総額の関係を保持

商品名抽出の最重要ルール【絶対遵守】:
- 送料関連項目でも、必ず元の意味のある名称を商品名として記録
- 「小計」「合計」「税込」「税抜」等の計算用語は商品名として絶対に使用禁止
- カタカナ表記の送料名称は、カタカナのまま保持（変換しない）
- レシート上の元の表記を最優先で使用する

【特別処理】「小計」表示の対応:
- 画像内で「チュウコカメラ ヨヤクソウリョウ」等の送料名称が「小計」と併記されている場合
- 必ず元の送料名称「チュウコカメラ ヨヤクソウリョウ」を商品名として使用
- 「小計」という語は完全に無視し、元の意味のある名称のみを抽出
- 例: 画像に「チュウコカメラ ヨヤクソウリョウ 小計 1650円」→ name: "チュウコカメラ ヨヤクソウリョウ"

店舗タイプ判別:
- Yahoo!オークション、ヤフオク → "ヤフオク" 
- Amazon、楽天 → "ECサイト"
- メルカリ、ラクマ → "フリマアプリ"
- 実店舗住所あり → "実店舗"
- その他 → "その他"

現在日時: {datetime.now().strftime('%Y/%m/%d')}（日付不明時の参考）

送料識別キーワード:
- 送料、配送料、配送費、運賃、運送料、小計（送料関連の場合）
- ソウリョウ、ハイソウリョウ、ハイソウヒ、ウンチン、ウンソウリョウ
- ヨヤクソウリョウ、チュウコカメラ ヨヤクソウリョウ、予約送料
- shipping、delivery、postage
- 宅配便、ゆうパック、レターパック、タクハイビン
- 佐川急便送料、ヤマト運輸送料、サガワキュウビン、ヤマトウンユ 等

JSON形式で出力:"""
                            },
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{base64_image}",
                                    "detail": "high"
                                }
                            }
                        ]
                    }
                ],
                max_tokens=6000,
                temperature=0.1
            )
            
            content = response.choices[0].message.content
            print(f"GPT-4o応答を受信: {len(content)}文字")
            
            # JSON抽出
            json_match = re.search(r'\{.*\}', content, re.DOTALL)
            if json_match:
                result = json.loads(json_match.group())
                result['file_name'] = image_path.name
                
                # インボイス番号の検証
                if result.get('invoice_number'):
                    validated_invoice = self.validate_invoice_number(result['invoice_number'])
                    result['invoice_number'] = validated_invoice
                    if validated_invoice:
                        print(f"✅ 有効なインボイス番号: {validated_invoice}")
                    else:
                        print(f"❌ インボイス番号が無効のため除外")
                
                # 住所フォーマット検証を追加
                if result.get('shop_address'):
                    is_valid, reason = self.validate_address_format(result['shop_address'])
                    if not is_valid:
                        print(f"🏠 住所除外: {result['shop_address']} → {reason}")
                        result['shop_address'] = ''
                    else:
                        print(f"✅ 住所有効: {result['shop_address']} ({reason})")
                
                # 住所が空の場合はAI検索で補完
                if not result.get('shop_address') and result.get('shop_name'):
                    print(f"🔍 住所が空のため、AI検索を実行: {result['shop_name']}")
                    ai_search_result = self.search_shop_address_with_ai(result['shop_name'])
                    
                    if ai_search_result.get('has_data') and ai_search_result.get('address'):
                        result['shop_address'] = ai_search_result['address']
                        print(f"✅ AI検索で住所を補完: {ai_search_result['address']}")
                        
                        # AI検索結果をメタデータとして記録
                        result['ai_search_info'] = {
                            'confidence': ai_search_result.get('confidence'),
                            'note': ai_search_result.get('note'),
                            'method': 'knowledge_cutoff_search'
                        }
                    else:
                        print(f"❌ AI検索でも住所が見つかりませんでした: {ai_search_result.get('note', 'N/A')}")
                
                return result
            else:
                raise Exception("JSONが見つかりませんでした")
                
        except Exception as e:
            print(f"解析エラー ({image_path.name}): {e}")
            return {
                'purchase_date': datetime.now().strftime('%Y/%m/%d'),  # 現在日付をデフォルト設定
                'products': [{'name': f'解析エラー: {image_path.name}', 'price': '0', 'quantity': 1}],
                'total': '0',
                'shop_name': '解析エラー',
                'shop_address': '',
                'payment_method': '',
                'shop_type': '',
                'invoice_number': '',
                'file_name': image_path.name,
                'error': f'解析エラー: {str(e)}'
            }

    def create_annual_summary_sheet(self, wb, year: int, months_with_data: List[int]):
        """年間合計シートを作成"""
        ws = wb.create_sheet("年間合計", 0)  # 最初のシートとして配置
        
        # スタイル設定
        header_font = Font(name='ＭＳ Ｐゴシック', bold=True, size=10)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # ヘッダー
        headers = ['月', '商品金額合計', '送料合計', '総合計']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 各月のデータ行
        row_num = 2
        for month in sorted(months_with_data):
            ws.cell(row=row_num, column=1, value=f"{month}月")
            
            # Excel数式で月別シートから集計
            sheet_name = f"{month}月"
            ws.cell(row=row_num, column=2, value=f"=SUM('{sheet_name}'!C:C)")  # 商品金額合計
            ws.cell(row=row_num, column=3, value=f"=SUM('{sheet_name}'!D:D)")  # 送料合計
            ws.cell(row=row_num, column=4, value=f"=SUM(B{row_num}:C{row_num})")  # 総合計
            
            row_num += 1
        
        # 年間合計行
        if row_num > 2:
            ws.cell(row=row_num + 1, column=1, value="年間合計")
            ws.cell(row=row_num + 1, column=2, value=f"=SUM(B2:B{row_num-1})")
            ws.cell(row=row_num + 1, column=3, value=f"=SUM(C2:C{row_num-1})")
            ws.cell(row=row_num + 1, column=4, value=f"=SUM(B{row_num+1}:C{row_num+1})")

    def create_excel_file(self, year: int, receipts: List[Dict[str, Any]]):
        """Excel仕入台帳を作成または更新（完成見本対応版）"""
        excel_path = self.output_folder / f"仕入台帳{year}.xlsx"
        
        # 既存ファイルがあれば読み込み、なければ新規作成
        if excel_path.exists():
            wb = openpyxl.load_workbook(excel_path)
        else:
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # スタイル設定
        header_font = Font(name='ＭＳ Ｐゴシック', bold=True, size=10)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # ヘッダー（完成見本に合わせる）
        headers = ['日付', '商品名', '金額', '送料', '合計金額', '店舗名', 
                   '店舗住所', '支払方法', '店舗タイプ', 'インボイス番号', '証憑ファイル']
        
        # 商品別行データの作成（送料特別処理対応）
        expanded_data = []
        current_month = datetime.now().month
        
        for receipt in receipts:
            target_month = current_month  # デフォルトは現在月
            
            # 購入日から月を抽出
            if receipt.get('purchase_date'):
                try:
                    date_parts = receipt['purchase_date'].split('/')
                    if len(date_parts) >= 2:
                        month = int(date_parts[1])
                        if 1 <= month <= 12:
                            target_month = month
                except:
                    pass
            
            # 商品ごとに行を作成（送料特別処理）
            if receipt.get('products') and len(receipt['products']) > 0:
                regular_shipping_fee = receipt.get('shipping_fee', '0')  # 別途記載の送料
                
                for product in receipt['products']:
                    is_shipping_item = product.get('is_shipping', False)
                    
                    if is_shipping_item:
                        # 送料項目の特別処理
                        row_data = {
                            'purchase_date': receipt.get('purchase_date', ''),
                            'product_name': product.get('name', '送料'),
                            'item_price': 0,  # 商品代金は0（数値型）
                            'shipping_fee': self._to_number(product.get('price', '0')),  # 送料欄に金額（数値型）
                            'shop_name': receipt.get('shop_name', ''),
                            'shop_address': receipt.get('shop_address', ''),
                            'payment_method': receipt.get('payment_method', ''),
                            'shop_type': receipt.get('shop_type', ''),
                            'invoice_number': receipt.get('invoice_number', ''),
                            'file_name': receipt.get('file_name', ''),
                            'month': target_month
                        }
                    else:
                        # 通常商品の処理
                        # 送料が商品リスト内にある場合は、通常商品には送料を付与しない
                        has_shipping_in_products = any(p.get('is_shipping', False) for p in receipt['products'])
                        
                        if has_shipping_in_products:
                            # 商品リスト内に送料あり → 通常商品には送料0
                            shipping_fee = 0
                        else:
                            # 商品リスト内に送料なし → 最初の商品に別途送料を付与
                            is_first_product = product == receipt['products'][0]
                            shipping_fee = self._to_number(regular_shipping_fee) if is_first_product else 0
                        
                        row_data = {
                            'purchase_date': receipt.get('purchase_date', ''),
                            'product_name': product.get('name', ''),
                            'item_price': self._to_number(product.get('price', '0')),  # 数値型
                            'shipping_fee': shipping_fee,  # 数値型
                            'shop_name': receipt.get('shop_name', ''),
                            'shop_address': receipt.get('shop_address', ''),
                            'payment_method': receipt.get('payment_method', ''),
                            'shop_type': receipt.get('shop_type', ''),
                            'invoice_number': receipt.get('invoice_number', ''),
                            'file_name': receipt.get('file_name', ''),
                            'month': target_month
                        }
                    
                    expanded_data.append(row_data)
            else:
                # 商品情報がない場合のフォールバック
                row_data = {
                    'purchase_date': receipt.get('purchase_date', ''),
                    'product_name': '商品情報なし',
                    'item_price': 0,
                    'shipping_fee': self._to_number(receipt.get('shipping_fee', '0')),
                    'shop_name': receipt.get('shop_name', ''),
                    'shop_address': receipt.get('shop_address', ''),
                    'payment_method': receipt.get('payment_method', ''),
                    'shop_type': receipt.get('shop_type', ''),
                    'invoice_number': receipt.get('invoice_number', ''),
                    'file_name': receipt.get('file_name', ''),
                    'month': target_month
                }
                expanded_data.append(row_data)
        
        # 月別にデータを整理
        monthly_data = {month: [] for month in range(1, 13)}
        for row in expanded_data:
            monthly_data[row['month']].append(row)
        
        # データのある月のシートを作成
        months_with_data = [month for month in range(1, 13) if monthly_data[month]]
        if not months_with_data:
            months_with_data = [current_month]
        
        for month in months_with_data:
            sheet_name = f"{month}月"
            
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
            else:
                ws = wb[sheet_name]
            
            # ヘッダー行を設定
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # データ行を設定
            row_num = 2
            for row_data in monthly_data[month]:
                # Excel数式での合計金額計算
                total_formula = f"=SUM(C{row_num}:D{row_num})"
                
                excel_row = [
                    row_data['purchase_date'],
                    row_data['product_name'],
                    row_data['item_price'],  # 数値型
                    row_data['shipping_fee'] if row_data['shipping_fee'] != 0 else None,  # 数値型またはNone
                    total_formula,  # Excel数式
                    row_data['shop_name'],
                    row_data['shop_address'],
                    row_data['payment_method'],
                    row_data['shop_type'],
                    row_data['invoice_number'],
                    row_data['file_name']
                ]
                
                for col, value in enumerate(excel_row, 1):
                    ws.cell(row=row_num, column=col, value=value)
                
                row_num += 1
        
        # 年間合計シートを作成
        self.create_annual_summary_sheet(wb, year, months_with_data)
        
        # ファイルを保存
        wb.save(excel_path)
        print(f"\n📊 Excel台帳を保存しました: {excel_path}")
        print(f"📝 {len(expanded_data)}行のデータを出力（商品別行分割・完成見本対応）")

    def _to_number(self, value) -> int:
        """文字列を数値に変換"""
        if value is None or value == '':
            return 0
        try:
            return int(str(value).replace(',', ''))
        except:
            return 0

    def run(self):
        """メイン処理"""
        print("=" * 60)
        print("🧾 証憑画像認識システム v1.3統合版(複数商品対応)")
        print("=" * 60)
        
        # APIキーの設定
        if not self.setup_api_key():
            print("\n❌ APIキーの設定に失敗しました")
            input("\nEnterで終了...")
            return
        
        # 新しい画像を取得
        new_images = self.get_new_images()
        
        if not new_images:
            print("\n📭 新しい画像が見つかりませんでした")
            print("   📁画像フォルダに画像を入れてください")
            input("\nEnterで終了...")
            return
        
        print(f"\n📸 {len(new_images)}個の新しい画像を処理します")
        
        # 画像を順次処理
        results = []
        processed_files = []
        
        for idx, image_path in enumerate(new_images, 1):
            print(f"\n[{idx}/{len(new_images)}] {image_path.name} を処理中...")
            
            # レシート解析
            result = self.analyze_receipt_with_vision(image_path)
            
            if result.get('error'):
                print(f"⚠️ エラーが発生: {result['error']}")
                continue
            
            # 結果表示
            print(f"📅 購入日: {result.get('purchase_date', 'N/A')}")
            if result.get('products'):
                print(f"🛍️ 商品数: {len(result['products'])}個")
                for product in result['products'][:3]:  # 最初の3個だけ表示
                    print(f"  - {product.get('name', 'N/A')}: ¥{product.get('price', '0')}")
            print(f"🏪 店舗名: {result.get('shop_name', 'N/A')}")
            
            results.append(result)
            
            # 処理済みとしてマーク
            image_hash = self.get_image_hash(image_path)
            if image_hash:
                self.processed_images.add(image_hash)
                processed_files.append(image_path)
        
        if not results:
            print("\n⚠️ 処理できた画像がありませんでした")
            input("\nEnterで終了...")
            return
        
        # 処理済みファイルリストを保存
        self.save_processed_images()
        
        # 年別にExcelファイルを作成
        yearly_data = {}
        current_year = datetime.now().year
        
        for result in results:
            year = current_year  # デフォルトは現在年
            
            if result.get('purchase_date'):
                try:
                    date_parts = result['purchase_date'].split('/')
                    if len(date_parts) >= 1:
                        year_candidate = int(date_parts[0])
                        if 1980 <= year_candidate <= current_year + 10:
                            year = year_candidate
                except:
                    pass
            
            if year not in yearly_data:
                yearly_data[year] = []
            yearly_data[year].append(result)
        
        # Excel作成
        for year, year_results in yearly_data.items():
            print(f"\n📊 {year}年のExcelファイルを作成中...")
            self.create_excel_file(year, year_results)
        
        # 処理済み画像をアーカイブフォルダに移動
        self.archive_processed_images(processed_files, results)
        
        print("\n" + "=" * 60)
        print("🎉 処理が完了しました！")
        print(f"📂 結果ファイル: {self.output_folder}")
        print("=" * 60)
        
        input("\nEnterで終了...")

if __name__ == "__main__":
    try:
        processor = PortableReceiptProcessorV3Multi()
        processor.run()
    except Exception as e:
        print(f"\n❌ システムエラー: {e}")
        import traceback
        traceback.print_exc()
        input("\nEnterで終了...")
